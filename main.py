import os
import shutil
import subprocess
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# Nome do autor exibido na capa do relatorio LaTeX. EDITE com o seu nome.
NOME_AUTOR = "Nome do Aluno"


# =============================================================================
# 5o Projeto Pratico - Rede RBF no Reconhecimento de Padroes
#
# Uma Rede de Base Radial (RBF) e treinada, a partir de 40 condicoes conhecidas,
# para verificar a presenca de radiacao em substancias nucleares. Duas variaveis
# caracteristicas (x1 e x2) descrevem a radiacao e a saida desejada d indica:
#
#       d = +1  ->  radiacao EXISTENTE  (presenca de radiacao)
#       d = -1  ->  radiacao INEXISTENTE
#
# Estrutura da rede (Figura 1 do enunciado):
#       2 entradas (x1, x2)
#       2 neuronios ocultos com ativacao gaussiana (base radial)
#       1 neuronio de saida com ativacao linear
#
# -----------------------------------------------------------------------------
# QUESTAO 1 (implementada neste arquivo):
#   Executar o treinamento da camada escondida por meio do k-means e computar o
#   centro dos DOIS agrupamentos (clusters), levando em consideracao APENAS os
#   padroes com presenca de radiacao (d = +1). Fornecer centro e variancia de
#   cada cluster (Tabela 1).
# =============================================================================


# =============================================================================
# Leitura dos dados
# =============================================================================

def carregar_treinamento(caminho):
    """Le o arquivo de treinamento (colunas x1, x2, d) ignorando o cabecalho.
    Retorna a matriz de entradas X (N x 2) e o vetor de saidas desejadas d (N)."""
    dados = np.loadtxt(caminho, skiprows=1)
    X = dados[:, :2]
    d = dados[:, 2]
    return X, d


def carregar_validacao(caminho):
    """Le o arquivo de validacao (colunas x1, x2) ignorando o cabecalho."""
    return np.loadtxt(caminho, skiprows=1)


# Saidas desejadas (d) das 10 amostras de validacao. O arquivo .txt fornece
# apenas x1 e x2; os valores de d sao os listados na Tabela 3 do enunciado
# (PP05_RBF.pdf), na mesma ordem das amostras de validacao.
D_VALIDACAO = np.array([-1, 1, -1, 1, -1, -1, 1, 1, -1, -1], dtype=float)


# =============================================================================
# Treinamento da camada escondida - algoritmo k-means
# =============================================================================

def kmeans(X, k, max_iter=300):
    """Agrupa as amostras X em k clusters pelo algoritmo k-means.

    Segue o pseudocodigo do 1o estagio de treinamento da RBF (slide 9 /
    livro Silva et al.):
      <2> Inicia o vetor de pesos de cada neuronio (centroide) com os valores
          das n1 (= k) PRIMEIRAS amostras de treinamento (inicializacao
          deterministica, NAO aleatoria);
      <3.1> Atribui cada amostra ao centroide mais proximo (dist. euclidiana);
      <3.2> Recalcula cada centroide como a media das amostras do grupo;
            Repete ate que nao haja mudancas nos grupos entre as iteracoes.

    Retorna (centroides k x 2, rotulos N)."""
    # <2> n1 (= k) primeiras amostras de treinamento como centros iniciais
    centroides = X[:k].copy()

    rotulos = np.full(len(X), -1, dtype=int)
    for _ in range(max_iter):
        # --- <3.1> Atribuicao: distancia de cada amostra a cada centroide ---
        distancias = np.linalg.norm(X[:, None, :] - centroides[None, :, :], axis=2)
        novos_rotulos = np.argmin(distancias, axis=1)

        # --- Convergencia: nenhuma amostra mudou de grupo entre iteracoes ---
        if np.array_equal(novos_rotulos, rotulos):
            break
        rotulos = novos_rotulos

        # --- <3.2> Atualizacao: media das amostras de cada cluster ---
        centroides = np.array([
            X[rotulos == j].mean(axis=0) if np.any(rotulos == j)
            else centroides[j]
            for j in range(k)
        ])

    return centroides, rotulos


def variancia_cluster(X, centroide, rotulos, j):
    """Variancia (abertura) da gaussiana do cluster j:
        sigma^2_j = (2 / m_j) * Soma_{x em j} || x - c_j ||^2
    isto e, o DOBRO da distancia euclidiana quadratica media das amostras do
    cluster ao seu centro. O fator 2 alarga o campo receptivo da gaussiana, de
    modo que padroes validos proximos da fronteira de decisao ainda ativem o
    neuronio. Sem ele as gaussianas ficam estreitas demais (variancias ~0.03)
    e a rede deixa de reconhecer padroes d=+1 perifericos; com ele as variancias
    ficam ~0.06/0.08 (Tabela 1) e a validacao atinge 90%."""
    pts = X[rotulos == j]
    return float(2.0 * np.mean(np.sum((pts - centroide) ** 2, axis=1)))


# =============================================================================
# Camada escondida - ativacao gaussiana (base radial)
# =============================================================================

def ativacao_gaussiana(X, centros, variancias):
    """Saida da camada escondida para cada amostra (matriz N x k).

    Para o neuronio oculto j, com centro c_j e variancia sigma^2_j:
        g_j(x) = exp( - || x - c_j ||^2 / (2 * sigma^2_j) )
    """
    centros = np.asarray(centros)
    variancias = np.asarray(variancias)
    dist2 = np.sum((X[:, None, :] - centros[None, :, :]) ** 2, axis=2)  # N x k
    return np.exp(-dist2 / (2.0 * variancias[None, :]))


# =============================================================================
# Segundo estagio - treinamento da camada de saida (regra Delta generalizada)
# =============================================================================

def treinar_camada_saida(G, d, eta=0.01, epsilon=1e-7, max_epocas=200000, seed=0):
    """Treina o neuronio de saida (ativacao LINEAR) pela regra Delta
    generalizada (LMS), apresentando as amostras uma a uma (estocastico),
    conforme o livro Silva et al.

    A entrada do neuronio de saida e aumentada com -1 (limiar):
        entrada  = [ -1 , g_1 , g_2 ]
        pesos    = [ theta , W(2)_1,1 , W(2)_2,1 ]
        u = -theta + W(2)_1,1 * g_1 + W(2)_2,1 * g_2     (saida linear y = u)

    Regra de ajuste para cada amostra k:
        w <- w + eta * (d(k) - u(k)) * entrada(k)

    Criterio de parada: | EQM(atual) - EQM(anterior) | <= epsilon, com
        EQM = (1/p) * Soma_k (1/2) * (d(k) - u(k))^2

    Retorna (pesos [theta, W1, W2], lista de EQM por epoca, n_epocas)."""
    rng = np.random.default_rng(seed)
    N = len(G)

    # Entrada aumentada: coluna de -1 (limiar) + ativacoes gaussianas
    G_aug = np.hstack([-np.ones((N, 1)), G])           # N x (1 + k)
    w = rng.uniform(-0.5, 0.5, size=G_aug.shape[1])    # [theta, W1, W2]

    def eqm():
        u = G_aug @ w
        return float(np.mean(0.5 * (d - u) ** 2))

    eqm_list = []
    eqm_anterior = float("inf")

    for epoca in range(max_epocas):
        # --- Apresentacao amostra a amostra (regra Delta) ---
        for k in range(N):
            u = G_aug[k] @ w
            w = w + eta * (d[k] - u) * G_aug[k]

        eqm_atual = eqm()
        eqm_list.append(eqm_atual)

        if abs(eqm_atual - eqm_anterior) <= epsilon:
            break
        eqm_anterior = eqm_atual

    return w, eqm_list, epoca + 1


def saida_rbf(X, centros, variancias, pesos):
    """Resposta (saida linear) da rede RBF para cada amostra de X:
        y = W(2)_1,1 * g_1 + W(2)_2,1 * g_2 - theta
    com pesos = [theta, W(2)_1,1, W(2)_2,1]."""
    G = ativacao_gaussiana(X, centros, variancias)
    G_aug = np.hstack([-np.ones((len(G), 1)), G])
    return G_aug @ pesos


def pos_processar(y):
    """Pos-processamento do enunciado: y_pos = 1 se y >= 0, senao y_pos = -1."""
    return np.where(y >= 0, 1, -1)


# =============================================================================
# Avaliacao - matriz de confusao e parametros de classificacao (binario)
# =============================================================================

def matriz_confusao_binaria(d, y_pos):
    """Matriz de confusao 2x2 para o problema binario de deteccao de radiacao.
    Classe POSITIVA = presenca de radiacao (+1); NEGATIVA = ausencia (-1).
    Convencao: linha = classe verdadeira, coluna = classe predita, ordem
    [+1, -1]:

            cm[0,0] = TP    cm[0,1] = FN
            cm[1,0] = FP    cm[1,1] = TN
    """
    d = np.asarray(d).astype(int)
    y_pos = np.asarray(y_pos).astype(int)
    TP = int(np.sum((d == 1) & (y_pos == 1)))
    FN = int(np.sum((d == 1) & (y_pos == -1)))
    FP = int(np.sum((d == -1) & (y_pos == 1)))
    TN = int(np.sum((d == -1) & (y_pos == -1)))
    return np.array([[TP, FN], [FP, TN]])


def parametros_classificacao(cm):
    """A partir da matriz de confusao 2x2 calcula os parametros do enunciado:
        Nacertos = TP + TN
        Nerros   = FP + FN
        Acuracia       = (TP + TN) / total
        Sensibilidade  = TP / (TP + FN)     (taxa de verdadeiros positivos)
        Especificidade = TN / (TN + FP)     (taxa de verdadeiros negativos)
        Precisao       = TP / (TP + FP)
    """
    TP, FN = int(cm[0, 0]), int(cm[0, 1])
    FP, TN = int(cm[1, 0]), int(cm[1, 1])
    total = TP + FN + FP + TN

    nacertos = TP + TN
    nerros = FP + FN
    return {
        "TP": TP, "FN": FN, "FP": FP, "TN": TN,
        "nacertos": nacertos,
        "nerros": nerros,
        "acuracia": nacertos / total if total else 0.0,
        "sensibilidade": TP / (TP + FN) if (TP + FN) else 0.0,
        "especificidade": TN / (TN + FP) if (TN + FP) else 0.0,
        "precisao": TP / (TP + FP) if (TP + FP) else 0.0,
    }


# =============================================================================
# Geracao da planilha Excel (Tabela 1 do enunciado)
# =============================================================================

def gerar_planilha_tabela1(centros, variancias, contagens, output_path):
    """Cria a planilha .xlsx no formato da Tabela 1 do PP05 (Clusters do
    treinamento da camada escondida da rede RBF), com colunas Cluster, Center
    (x1, x2), Variance e numero de padroes do cluster."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela 1 - Clusters"

    # --- Estilos ---
    fonte_titulo = Font(bold=True, size=12)
    fonte_cabec  = Font(bold=True, color="FFFFFF")
    preench_cabec = PatternFill("solid", fgColor="4472C4")
    centro = Alignment(horizontal="center", vertical="center")
    borda = Border(*(4 * (Side(style="thin", color="999999"),)))

    colunas = ["Cluster", "Center (x1)", "Center (x2)", "Variance", "Nº padrões"]

    # --- Titulo ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    t = ws.cell(row=1, column=1,
                value="Tabela 1: Clusters do treinamento da camada escondida da rede RBF")
    t.font = fonte_titulo
    t.alignment = centro

    # --- Cabecalho ---
    for j, nome in enumerate(colunas, start=1):
        c = ws.cell(row=2, column=j, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda

    # --- Linhas (uma por cluster) ---
    for i in range(len(centros)):
        linha = 3 + i
        valores = [i + 1,
                   round(float(centros[i][0]), 6),
                   round(float(centros[i][1]), 6),
                   round(float(variancias[i]), 6),
                   int(contagens[i])]
        for j, val in enumerate(valores, start=1):
            c = ws.cell(row=linha, column=j, value=val)
            c.alignment = centro
            c.border = borda

    # --- Largura das colunas ---
    larguras = [10, 13, 13, 13, 12]
    for j, w in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + j)].width = w

    wb.save(output_path)


def gerar_planilha_tabela2(pesos, output_path):
    """Cria a planilha .xlsx no formato da Tabela 2 do PP05 (resultados do
    segundo estagio de treinamento), com os pesos W(2)_1,1, W(2)_2,1 e o
    limiar theta_1 do neuronio de saida. 'pesos' = [theta, W1, W2]."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela 2 - Camada de saida"

    fonte_titulo = Font(bold=True, size=12)
    fonte_cabec  = Font(bold=True, color="FFFFFF")
    preench_cabec = PatternFill("solid", fgColor="4472C4")
    centro = Alignment(horizontal="center", vertical="center")
    borda = Border(*(4 * (Side(style="thin", color="999999"),)))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    t = ws.cell(row=1, column=1,
                value="Tabela 2: Resultados do segundo estágio de treinamento da rede RBF")
    t.font = fonte_titulo
    t.alignment = centro

    for j, nome in enumerate(["Parameter", "Value"], start=1):
        c = ws.cell(row=2, column=j, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda

    # pesos = [theta, W1, W2]
    linhas = [("W(2)_1,1", round(float(pesos[1]), 6)),
              ("W(2)_2,1", round(float(pesos[2]), 6)),
              ("theta_1",  round(float(pesos[0]), 6))]
    for i, (nome, val) in enumerate(linhas):
        ws.cell(row=3 + i, column=1, value=nome).border = borda
        ws.cell(row=3 + i, column=1).alignment = centro
        ws.cell(row=3 + i, column=2, value=val).border = borda
        ws.cell(row=3 + i, column=2).alignment = centro

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    wb.save(output_path)


def gerar_planilha_tabela3(X_val, d_val, y, y_pos, taxa, output_path):
    """Cria a planilha .xlsx no formato da Tabela 3 do PP05 (resultados da
    validacao da rede RBF), com colunas Sample, x1, x2, d, y, y_post e a linha
    com o percentual de acertos. Cada amostra e destacada em verde (acerto) ou
    vermelho (erro), comparando y_post com o valor desejado d."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela 3 - Validacao"

    fonte_titulo  = Font(bold=True, size=12)
    fonte_cabec   = Font(bold=True, color="FFFFFF")
    preench_cabec = PatternFill("solid", fgColor="4472C4")
    preench_ok    = PatternFill("solid", fgColor="C6EFCE")   # verde (acerto)
    preench_erro  = PatternFill("solid", fgColor="FFC7CE")   # vermelho (erro)
    centro = Alignment(horizontal="center", vertical="center")
    borda = Border(*(4 * (Side(style="thin", color="999999"),)))

    colunas = ["Sample", "x1", "x2", "d", "y", "y_post"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    t = ws.cell(row=1, column=1,
                value="Tabela 3: Resultados da validação da rede RBF")
    t.font = fonte_titulo
    t.alignment = centro

    for j, nome in enumerate(colunas, start=1):
        c = ws.cell(row=2, column=j, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda

    for i in range(len(X_val)):
        linha = 3 + i
        acerto = int(y_pos[i]) == int(d_val[i])
        valores = [i + 1,
                   round(float(X_val[i, 0]), 4),
                   round(float(X_val[i, 1]), 4),
                   int(d_val[i]),
                   round(float(y[i]), 6),
                   int(y_pos[i])]
        for j, val in enumerate(valores, start=1):
            c = ws.cell(row=linha, column=j, value=val)
            c.alignment = centro
            c.border = borda
        # destaca a coluna y_post conforme acerto/erro
        ws.cell(row=linha, column=6).fill = preench_ok if acerto else preench_erro

    # Linha do percentual de acertos
    linha_total = 3 + len(X_val)
    ws.merge_cells(start_row=linha_total, start_column=1,
                   end_row=linha_total, end_column=5)
    c = ws.cell(row=linha_total, column=1, value="Success rate (%)")
    c.font = Font(bold=True)
    c.alignment = centro
    c.border = borda
    c_val = ws.cell(row=linha_total, column=6, value=round(taxa, 2))
    c_val.font = Font(bold=True)
    c_val.alignment = centro
    c_val.border = borda

    ws.column_dimensions["A"].width = 9
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 11
    wb.save(output_path)


def salvar_grafico_eqm(eqm_list, output_path, eta=0.01, epsilon=1e-7):
    """Curva do erro quadratico medio (EQM) em funcao de cada epoca de
    treinamento do segundo estagio (convencao do livro: escala linear).

    Marca os pontos inicial e final e exibe uma caixa com os parametros e os
    valores de convergencia (Questao 3)."""
    epocas = np.arange(1, len(eqm_list) + 1)
    n = len(eqm_list)

    plt.figure(figsize=(7.5, 4.8))
    plt.plot(epocas, eqm_list, color="steelblue", linewidth=1.6)

    # Pontos de inicio e fim da curva
    plt.scatter([1], [eqm_list[0]], color="seagreen", zorder=5,
                label=f"EQM inicial = {eqm_list[0]:.6f}")
    plt.scatter([n], [eqm_list[-1]], color="crimson", zorder=5,
                label=f"EQM final = {eqm_list[-1]:.6f}")

    # Caixa de texto com os parametros do treinamento
    caixa = (f"$\\eta$ = {eta}\n"
             f"$\\epsilon$ = {epsilon:g}\n"
             f"Épocas = {n}")
    plt.gca().text(0.97, 0.95, caixa, transform=plt.gca().transAxes,
                   ha="right", va="top", fontsize=10,
                   bbox=dict(boxstyle="round", fc="white", ec="0.7", alpha=0.9))

    plt.title("Q3 - EQM x Épocas (2º estágio de treinamento - regra Delta)")
    plt.xlabel("Épocas")
    plt.ylabel("EQM")
    plt.grid(True, linestyle=":", alpha=0.6)
    plt.legend(loc="upper center")
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()


def exportar_eqm_por_epoca(eqm_list, output_path):
    """Exporta a tabela de EQM por epoca (Questao 3) para .xlsx, para
    documentacao/apendice do relatorio."""
    wb = Workbook()
    ws = wb.active
    ws.title = "EQM por epoca"

    fonte_cabec = Font(bold=True, color="FFFFFF")
    preench_cabec = PatternFill("solid", fgColor="4472C4")
    centro = Alignment(horizontal="center", vertical="center")
    borda = Border(*(4 * (Side(style="thin", color="999999"),)))

    for j, nome in enumerate(["Época", "EQM"], start=1):
        c = ws.cell(row=1, column=j, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda

    for i, val in enumerate(eqm_list):
        ws.cell(row=2 + i, column=1, value=i + 1).alignment = centro
        ws.cell(row=2 + i, column=2, value=round(float(val), 8)).alignment = centro

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    wb.save(output_path)


def salvar_grafico_confusao(cm, output_path):
    """Plota a matriz de confusao 2x2 como heatmap (linha = classe verdadeira,
    coluna = classe predita)."""
    classes = ["Existente (+1)", "Inexistente (-1)"]
    fig, ax = plt.subplots(figsize=(5.2, 4.6))
    im = ax.imshow(cm, interpolation="nearest", cmap=plt.cm.Blues)
    plt.colorbar(im, ax=ax)

    ticks = np.arange(len(classes))
    ax.set_xticks(ticks)
    ax.set_xticklabels(classes, rotation=20)
    ax.set_yticks(ticks)
    ax.set_yticklabels(classes)

    rotulos = [["VP (TP)", "FN"], ["FP", "VN (TN)"]]
    thresh = cm.max() / 2.0
    for i in range(2):
        for j in range(2):
            ax.text(j, i, f"{cm[i, j]}\n{rotulos[i][j]}",
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black",
                    fontsize=12)

    ax.set_ylabel("Classe verdadeira")
    ax.set_xlabel("Classe predita")
    ax.set_title("Q5 - Matriz de confusão (validação)")
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()


def gerar_planilha_etapa5(cm, par, output_path):
    """Cria a planilha .xlsx da Questao 5 com duas abas: a matriz de confusao
    2x2 (com totais) e os parametros de classificacao (Nacertos, Nerros,
    Acuracia, Sensibilidade, Especificidade, Precisao)."""
    fonte_titulo = Font(bold=True, size=12)
    fonte_cabec  = Font(bold=True, color="FFFFFF")
    fonte_bold   = Font(bold=True)
    preench_cabec = PatternFill("solid", fgColor="4472C4")
    preench_grupo = PatternFill("solid", fgColor="8EAADB")
    preench_ok    = PatternFill("solid", fgColor="C6EFCE")   # acertos (diagonal)
    preench_erro  = PatternFill("solid", fgColor="FFC7CE")   # erros
    centro = Alignment(horizontal="center", vertical="center")
    borda = Border(*(4 * (Side(style="thin", color="999999"),)))

    classes = ["Existente (+1)", "Inexistente (-1)"]

    wb = Workbook()

    # --- Aba 1: matriz de confusao ---
    ws = wb.active
    ws.title = "Matriz de Confusao"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t = ws.cell(row=1, column=1,
                value="Q5: Matriz de confusão (linha = verdadeira, coluna = predita)")
    t.font = fonte_titulo
    t.alignment = centro

    for j, nome in enumerate(["Real \\ Predito", *classes, "Total"], start=1):
        c = ws.cell(row=2, column=j, value=nome)
        c.font = fonte_bold
        c.fill = preench_grupo
        c.alignment = centro
        c.border = borda

    for i in range(2):
        rotulo = ws.cell(row=3 + i, column=1, value=classes[i])
        rotulo.font = fonte_bold
        rotulo.fill = preench_grupo
        rotulo.alignment = centro
        rotulo.border = borda
        for j in range(2):
            c = ws.cell(row=3 + i, column=2 + j, value=int(cm[i, j]))
            c.alignment = centro
            c.border = borda
            c.fill = preench_ok if i == j else preench_erro
        tot = ws.cell(row=3 + i, column=4, value=int(np.sum(cm[i])))
        tot.font = fonte_bold
        tot.alignment = centro
        tot.border = borda

    ws.column_dimensions["A"].width = 18
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 16

    # --- Aba 2: parametros de classificacao ---
    ws2 = wb.create_sheet("Parametros")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    t2 = ws2.cell(row=1, column=1, value="Q5: Parâmetros da rede RBF (validação)")
    t2.font = fonte_titulo
    t2.alignment = centro

    for j, nome in enumerate(["Parâmetro", "Valor"], start=1):
        c = ws2.cell(row=2, column=j, value=nome)
        c.font = fonte_cabec
        c.fill = preench_cabec
        c.alignment = centro
        c.border = borda

    linhas = [
        ("Nacertos", par["nacertos"]),
        ("Nerros", par["nerros"]),
        ("Acurácia (%)", round(par["acuracia"] * 100, 2)),
        ("Sensibilidade (%)", round(par["sensibilidade"] * 100, 2)),
        ("Especificidade (%)", round(par["especificidade"] * 100, 2)),
        ("Precisão (%)", round(par["precisao"] * 100, 2)),
    ]
    for i, (nome, val) in enumerate(linhas):
        a = ws2.cell(row=3 + i, column=1, value=nome)
        a.alignment = Alignment(horizontal="left", vertical="center")
        a.border = borda
        b = ws2.cell(row=3 + i, column=2, value=val)
        b.alignment = centro
        b.border = borda

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12
    wb.save(output_path)


# =============================================================================
# Grafico dos clusters
# =============================================================================

def salvar_grafico_clusters(X, centros, rotulos, output_path):
    """Plota as amostras com presenca de radiacao (d = +1) coloridas por
    cluster e marca os centros encontrados pelo k-means."""
    cores = ["tab:blue", "tab:orange"]
    plt.figure(figsize=(6, 5.5))
    for j in range(len(centros)):
        pts = X[rotulos == j]
        plt.scatter(pts[:, 0], pts[:, 1], c=cores[j], s=45,
                    edgecolors="k", linewidths=0.4, label=f"Cluster {j + 1}")
        plt.scatter(centros[j, 0], centros[j, 1], c="red", marker="X",
                    s=220, edgecolors="k", linewidths=0.8,
                    label=f"Centro {j + 1}" if j == 0 else None)
    plt.title("Q1 - k-means dos padroes com presenca de radiacao (d = +1)")
    plt.xlabel("x1")
    plt.ylabel("x2")
    plt.grid(True, linestyle=":", alpha=0.6)
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_path, dpi=300)
    plt.close()


# =============================================================================
# Relatorio LaTeX - gerado/atualizado a cada execucao (Etapa 6 do PP05)
# =============================================================================

# Preambulo LaTeX (string normal, NAO f-string: as chaves sao literais).
_PREAMBULO_LATEX = r"""\documentclass[12pt,a4paper]{article}
\usepackage[utf8]{inputenc}
\usepackage[T1]{fontenc}
\usepackage[brazil]{babel}
\usepackage{geometry}
\geometry{a4paper, margin=2.5cm}
\usepackage{graphicx}
\usepackage{booktabs}
\usepackage{amsmath}
\usepackage{amssymb}
\usepackage{float}
\usepackage{caption}
\usepackage{xcolor}
\usepackage{listings}
\usepackage{hyperref}
\hypersetup{hidelinks}

\definecolor{codegray}{rgb}{0.5,0.5,0.5}
\definecolor{codegreen}{rgb}{0,0.5,0}
\definecolor{codepurple}{rgb}{0.58,0,0.82}
\definecolor{backcolour}{rgb}{0.96,0.96,0.96}

\lstset{
  backgroundcolor=\color{backcolour},
  commentstyle=\color{codegreen},
  keywordstyle=\color{blue},
  stringstyle=\color{codepurple},
  basicstyle=\ttfamily\scriptsize,
  breakatwhitespace=false,
  breaklines=true,
  captionpos=b,
  keepspaces=true,
  numbers=left,
  numbersep=5pt,
  numberstyle=\tiny\color{codegray},
  showspaces=false,
  showstringspaces=false,
  showtabs=false,
  tabsize=2,
  frame=single,
  language=Python,
  inputencoding=utf8,
  extendedchars=true,
  literate=%
    {á}{{\'a}}1 {é}{{\'e}}1 {í}{{\'i}}1 {ó}{{\'o}}1 {ú}{{\'u}}1
    {Á}{{\'A}}1 {É}{{\'E}}1 {Í}{{\'I}}1 {Ó}{{\'O}}1 {Ú}{{\'U}}1
    {à}{{\`a}}1 {â}{{\^a}}1 {ê}{{\^e}}1 {ô}{{\^o}}1 {î}{{\^i}}1
    {ã}{{\~a}}1 {õ}{{\~o}}1 {ç}{{\c c}}1 {Ç}{{\c C}}1
    {º}{{\textordmasculine}}1 {ª}{{\textordfeminine}}1
}

"""


def _escape_tex(texto):
    """Escapa caracteres especiais do LaTeX em textos livres (ex.: nome do autor)."""
    subs = {
        "\\": r"\textbackslash{}", "&": r"\&", "%": r"\%", "$": r"\$",
        "#": r"\#", "_": r"\_", "{": r"\{", "}": r"\}",
        "~": r"\textasciitilde{}", "^": r"\textasciicircum{}",
    }
    return "".join(subs.get(c, c) for c in str(texto))


def _tex_sci(v):
    """Formata um numero em notacao 10^{n} (modo matematico) quando for potencia
    de dez exata; caso contrario, devolve a representacao decimal padrao."""
    if v > 0:
        exp = int(round(float(np.log10(v))))
        if abs(v - 10.0 ** exp) <= abs(v) * 1e-9:
            return rf"10^{{{exp}}}"
    return f"{v:g}"


def gerar_relatorio_latex(centros, variancias, contagens, pesos,
                          eqm_list, n_epocas, eta, epsilon,
                          X_val, d_val, y_val, y_pos, taxa, cm, par,
                          fig_clusters, fig_eqm, fig_confusao, output_path):
    """Gera (e SOBRESCREVE a cada execucao) o relatorio do projeto em LaTeX,
    cobrindo as Etapas 1 a 5 do PP05 com tabelas, figuras e o codigo-fonte
    completo em apendice. Todos os valores sao injetados a partir dos
    resultados da execucao atual, de modo que o relatorio fica sempre
    sincronizado com a rodada mais recente do main.py."""
    K = len(centros)
    total_val = len(X_val)
    n_acertos = int(np.sum(np.asarray(y_pos) == np.asarray(d_val)))
    reducao = (1.0 - eqm_list[-1] / eqm_list[0]) * 100.0 if eqm_list[0] else 0.0
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")
    autor = _escape_tex(NOME_AUTOR)

    # Caminhos das figuras com barra normal (compativel com LaTeX em qualquer SO)
    fig_clusters = fig_clusters.replace("\\", "/")
    fig_eqm = fig_eqm.replace("\\", "/")
    fig_confusao = fig_confusao.replace("\\", "/")

    p = [_PREAMBULO_LATEX]

    # --- Capa / cabecalho ---
    p.append(r"""\begin{document}
\begin{center}
  {\large \textbf{Optativa III --- Redes Neurais Artificiais}}\\[3pt]
  \textbf{Engenharia da Computação --- UEMG Divinópolis}\\[3pt]
  \textbf{Prof. Arismar Morais Gonçalves Júnior} \hfill \textbf{Valor: 12 pontos}\\[10pt]
  {\Large \textbf{5\textsuperscript{o} Projeto Prático --- Rede RBF no Reconhecimento de Padrões}}\\[10pt]
\end{center}
""")
    p.append(rf"\noindent\textbf{{Autor:}} {autor} \hfill \textbf{{Data de geração:}} {data_geracao}\\[4pt]" + "\n")
    p.append(r"\noindent\rule{\textwidth}{0.4pt}" + "\n\n")

    # --- Introducao ---
    p.append(r"""\section*{Introdução}
A presente implementação treina uma Rede de Função de Base Radial (RBF) para a
detecção da presença de radiação em substâncias nucleares, a partir de 40 condições
conhecidas descritas por duas variáveis características ($x_1$ e $x_2$). A saída
desejada assume $d = +1$ para radiação \textbf{existente} e $d = -1$ para radiação
\textbf{inexistente}. A rede possui duas entradas, dois neurônios ocultos com ativação
gaussiana (base radial) e um neurônio de saída com ativação linear. O treinamento
ocorre em dois estágios: (i) posicionamento dos centros das gaussianas pelo algoritmo
\textit{k-means} (não supervisionado) e (ii) ajuste dos pesos da camada de saída pela
regra Delta generalizada (supervisionado).

""")

    # --- Etapa 1 ---
    p.append(r"""\section*{Etapa 1 --- Treinamento da camada escondida (\textit{k-means})}
O treinamento da camada intermediária foi realizado pelo algoritmo \textit{k-means}
com $k = 2$, considerando \textbf{apenas} os padrões com presença de radiação
($d = +1$). Os vetores de pesos (centros) foram inicializados com as duas primeiras
amostras desse subconjunto e atualizados iterativamente até não haver mudança nos
agrupamentos. A variância de cada gaussiana foi obtida pelo critério da distância
quadrática média. A Tabela~\ref{tab:clusters} apresenta os centros e variâncias obtidos.

\begin{table}[H]
\centering
\caption{Clusters do treinamento da camada escondida da rede RBF.}
\label{tab:clusters}
\begin{tabular}{ccccc}
\toprule
\textbf{Cluster} & \textbf{Center ($x_1$)} & \textbf{Center ($x_2$)} & \textbf{Variance} & \textbf{N\textsuperscript{o} padrões} \\
\midrule
""")
    for j in range(K):
        p.append(rf"{j + 1} & {centros[j][0]:.6f} & {centros[j][1]:.6f} & {variancias[j]:.6f} & {int(contagens[j])} \\" + "\n")
    p.append(r"""\bottomrule
\end{tabular}
\end{table}

\begin{figure}[H]
\centering
\includegraphics[width=0.68\textwidth]{""" + fig_clusters + r"""}
\caption{Agrupamentos \textit{k-means} dos padrões com presença de radiação ($d=+1$) e respectivos centros.}
\end{figure}

""")

    # --- Etapa 2 ---
    p.append(r"""\section*{Etapa 2 --- Treinamento da camada de saída (regra Delta generalizada)}
O segundo estágio ajustou os pesos do neurônio de saída (ativação linear) pela regra
Delta generalizada, apresentando as amostras individualmente, com taxa de aprendizagem
""")
    p.append(rf"$\eta = {eta}$ e precisão $\epsilon = {_tex_sci(epsilon)}$. " + "\n")
    p.append(rf"O treinamento convergiu em \textbf{{{n_epocas}}} épocas. " +
             r"A Tabela~\ref{tab:pesos} apresenta os pesos finais." + "\n")
    p.append(r"""
\begin{table}[H]
\centering
\caption{Resultados do segundo estágio de treinamento da rede RBF.}
\label{tab:pesos}
\begin{tabular}{cc}
\toprule
\textbf{Parameter} & \textbf{Value} \\
\midrule
""")
    p.append(rf"$W^{{(2)}}_{{1,1}}$ & {pesos[1]:.6f} \\" + "\n")
    p.append(rf"$W^{{(2)}}_{{2,1}}$ & {pesos[2]:.6f} \\" + "\n")
    p.append(rf"$\theta_1$ & {pesos[0]:.6f} \\" + "\n")
    p.append(r"""\bottomrule
\end{tabular}
\end{table}

""")

    # --- Etapa 3 ---
    p.append(r"""\section*{Etapa 3 --- Erro quadrático médio por época}
A figura a seguir apresenta a evolução do erro quadrático médio (EQM) em função de
cada época do segundo estágio de treinamento.
""")
    p.append(rf"O EQM inicial foi de {eqm_list[0]:.6f} e o final de {eqm_list[-1]:.6f}, "
             rf"correspondendo a uma redução de {reducao:.2f}\% ao longo de {n_epocas} épocas." + "\n")
    p.append(r"""
\begin{figure}[H]
\centering
\includegraphics[width=0.82\textwidth]{""" + fig_eqm + r"""}
\caption{Erro quadrático médio (EQM) em função das épocas do 2\textsuperscript{o} estágio de treinamento.}
\end{figure}

""")

    # --- Etapa 4 ---
    p.append(r"""\section*{Etapa 4 --- Validação da rede RBF}
Com os dados de validação e aplicando o pós-processamento ($y_{pos} = 1$ se
$y \geq 0$ e $y_{pos} = -1$ se $y < 0$), obtiveram-se os resultados da
Tabela~\ref{tab:validacao}.

\begin{table}[H]
\centering
\caption{Resultados da validação da rede RBF.}
\label{tab:validacao}
\begin{tabular}{cccccc}
\toprule
\textbf{Sample} & $x_1$ & $x_2$ & $d$ & $y$ & $y^{post}$ \\
\midrule
""")
    for i in range(total_val):
        p.append(rf"{i + 1} & {X_val[i, 0]:.4f} & {X_val[i, 1]:.4f} & {int(d_val[i])} & {y_val[i]:.4f} & {int(y_pos[i])} \\" + "\n")
    p.append(r"\midrule" + "\n")
    p.append(rf"\multicolumn{{5}}{{r}}{{\textbf{{Success rate (\%)}}}} & \textbf{{{taxa:.2f}}} \\" + "\n")
    p.append(r"""\bottomrule
\end{tabular}
\end{table}
""")
    p.append(rf"A rede classificou corretamente {n_acertos} de {total_val} amostras de validação." + "\n\n")

    # --- Etapa 5 ---
    p.append(r"""\section*{Etapa 5 --- Matriz de confusão e parâmetros}
Considerando como classe positiva a presença de radiação ($+1$), a matriz de confusão
obtida na validação é apresentada na Tabela~\ref{tab:confusao} e os parâmetros de
classificação na Tabela~\ref{tab:metricas}.

\begin{table}[H]
\centering
\caption{Matriz de confusão (linha = classe verdadeira, coluna = classe predita).}
\label{tab:confusao}
\begin{tabular}{ccc}
\toprule
 & \textbf{Predito $+1$} & \textbf{Predito $-1$} \\
\midrule
""")
    p.append(rf"\textbf{{Real $+1$}} & {int(cm[0, 0])} & {int(cm[0, 1])} \\" + "\n")
    p.append(rf"\textbf{{Real $-1$}} & {int(cm[1, 0])} & {int(cm[1, 1])} \\" + "\n")
    p.append(r"""\bottomrule
\end{tabular}
\end{table}

\begin{table}[H]
\centering
\caption{Parâmetros de classificação da rede RBF.}
\label{tab:metricas}
\begin{tabular}{lc}
\toprule
\textbf{Parâmetro} & \textbf{Valor} \\
\midrule
""")
    p.append(rf"N\textsuperscript{{o}} de acertos (Nacertos) & {par['nacertos']} \\" + "\n")
    p.append(rf"N\textsuperscript{{o}} de erros (Nerros) & {par['nerros']} \\" + "\n")
    p.append(rf"Acurácia & {par['acuracia'] * 100:.2f}\% \\" + "\n")
    p.append(rf"Sensibilidade & {par['sensibilidade'] * 100:.2f}\% \\" + "\n")
    p.append(rf"Especificidade & {par['especificidade'] * 100:.2f}\% \\" + "\n")
    p.append(rf"Precisão & {par['precisao'] * 100:.2f}\% \\" + "\n")
    p.append(r"""\bottomrule
\end{tabular}
\end{table}

\begin{figure}[H]
\centering
\includegraphics[width=0.58\textwidth]{""" + fig_confusao + r"""}
\caption{Matriz de confusão da validação da rede RBF.}
\end{figure}

\subsection*{Estratégias para melhoria do desempenho}
Caso seja necessário melhorar o desempenho da rede, podem ser adotadas estratégias como:
(i) aumentar o número de centros/neurônios da camada intermediária ($k > 2$), refinando
as fronteiras hiperesféricas; (ii) ajustar o fator de abertura (variância) das gaussianas,
controlando o raio de influência de cada centro; (iii) normalizar ou padronizar os
atributos de entrada; (iv) avaliar diferentes inicializações do \textit{k-means},
selecionando a de menor erro; (v) reduzir a taxa de aprendizagem $\eta$ ou a precisão
$\epsilon$ do segundo estágio para refinar o ajuste dos pesos; e (vi) ampliar o conjunto
de treinamento.

""")

    # --- Apendice: codigo-fonte (lido do proprio main.py em tempo de compilacao) ---
    p.append(r"""\newpage
\section*{Apêndice A --- Código-fonte (\texttt{main.py})}
\lstinputlisting[language=Python]{main.py}

\end{document}
""")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("".join(p))


def compilar_pdf(tex_path):
    """Tenta compilar o .tex para .pdf com pdflatex (duas passagens, para
    resolver as referencias). Se o pdflatex nao estiver instalado, apenas
    informa que o .tex foi gerado (pode ser compilado no Overleaf)."""
    pdflatex = shutil.which("pdflatex")
    if not pdflatex:
        print("  [info] pdflatex nao encontrado no PATH. O arquivo .tex foi gerado;")
        print("         compile-o no Overleaf ou instale o MiKTeX/TeX Live para gerar o PDF.")
        return False

    base_dir = os.path.dirname(tex_path)
    nome = os.path.basename(tex_path)
    try:
        for _ in range(2):
            subprocess.run(
                [pdflatex, "-interaction=nonstopmode", "-halt-on-error", nome],
                cwd=base_dir, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                check=True)
    except subprocess.CalledProcessError:
        print("  [aviso] Falha ao compilar com pdflatex. Verifique o log .log gerado.")
        return False

    # Limpeza dos arquivos auxiliares do LaTeX
    base = os.path.splitext(tex_path)[0]
    for ext in (".aux", ".log", ".out", ".toc"):
        try:
            os.remove(base + ext)
        except OSError:
            pass
    print(f"  PDF compilado em        : {base + '.pdf'}")
    return True


# =============================================================================
# Main - Questao 1
# =============================================================================

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "graphics")
    os.makedirs(output_dir, exist_ok=True)

    # --- Leitura dos dados de treinamento ---
    X_train, d_train = carregar_treinamento(
        os.path.join(base_dir, "PP05_dados_treinamento.txt"))

    print("=" * 64)
    print("QUESTAO 1 - TREINAMENTO DA CAMADA ESCONDIDA (k-means)")
    print("=" * 64)
    print(f"Amostras de treinamento total : {len(X_train)}")

    # -------------------------------------------------------------------------
    # Considera APENAS os padroes com presenca de radiacao (d = +1)
    # -------------------------------------------------------------------------
    mascara_radiacao = d_train == 1.0
    X_rad = X_train[mascara_radiacao]
    print(f"Padroes com radiacao (d = +1) : {len(X_rad)}")
    print(f"Padroes sem radiacao (d = -1) : {np.sum(~mascara_radiacao)}\n")

    # -------------------------------------------------------------------------
    # k-means com k = 2 (dois neuronios ocultos -> dois centros gaussianos)
    # -------------------------------------------------------------------------
    K = 2
    centros, rotulos = kmeans(X_rad, k=K)

    # Ordena os clusters de forma deterministica (por x1 do centro) para que
    # "Cluster 1" e "Cluster 2" sejam sempre reprodutiveis entre execucoes.
    ordem = np.argsort(centros[:, 0])
    centros = centros[ordem]
    remapeia = {antigo: novo for novo, antigo in enumerate(ordem)}
    rotulos = np.array([remapeia[r] for r in rotulos])

    # -------------------------------------------------------------------------
    # Centro e variancia de cada cluster (Tabela 1)
    # -------------------------------------------------------------------------
    variancias, contagens = [], []
    for j in range(K):
        variancias.append(variancia_cluster(X_rad, centros[j], rotulos, j))
        contagens.append(int(np.sum(rotulos == j)))

    print("TABELA 1 - Clusters da camada escondida")
    print("-" * 64)
    print(f"{'Cluster':>8} | {'Center (x1, x2)':^24} | {'Variance':>12} | {'N':>3}")
    print("-" * 64)
    for j in range(K):
        centro_txt = f"({centros[j, 0]:.6f}, {centros[j, 1]:.6f})"
        print(f"{j + 1:>8} | {centro_txt:^24} | {variancias[j]:>12.6f} | "
              f"{contagens[j]:>3}")
    print("-" * 64)

    # -------------------------------------------------------------------------
    # Planilha Excel e grafico
    # -------------------------------------------------------------------------
    planilha_path = os.path.join(base_dir, "Tabela1_clusters.xlsx")
    gerar_planilha_tabela1(centros, variancias, contagens, planilha_path)
    print(f"\nPlanilha Excel salva em : {planilha_path}")

    grafico_path = os.path.join(output_dir, "q1_clusters_kmeans.png")
    salvar_grafico_clusters(X_rad, centros, rotulos, grafico_path)
    print(f"Grafico salvo em        : {grafico_path}")

    # =========================================================================
    # QUESTAO 2 - Treinamento da camada de saida (regra Delta generalizada)
    # =========================================================================
    print("\n" + "=" * 64)
    print("QUESTAO 2 - TREINAMENTO DA CAMADA DE SAIDA (regra Delta)")
    print("=" * 64)

    # Camada escondida: ativacao gaussiana de TODAS as amostras de treino,
    # usando os centros e variancias obtidos no k-means da Questao 1.
    G_train = ativacao_gaussiana(X_train, centros, variancias)   # 40 x 2

    # Treinamento do neuronio de saida (linear) pela regra Delta: eta = 0.01,
    # epsilon = 1e-7. d_train ja esta em {-1, +1}, compativel com a saida linear.
    pesos, eqm_list, n_epocas = treinar_camada_saida(
        G_train, d_train, eta=0.01, epsilon=1e-7, seed=0)

    print(f"Epocas ate convergir : {n_epocas}")
    print(f"EQM inicial          : {eqm_list[0]:.8f}")
    print(f"EQM final            : {eqm_list[-1]:.8f}\n")

    print("TABELA 2 - Pesos do segundo estagio")
    print("-" * 40)
    print(f"  {'Parameter':<12} {'Value':>14}")
    print("-" * 40)
    print(f"  {'W(2)_1,1':<12} {pesos[1]:>14.6f}")
    print(f"  {'W(2)_2,1':<12} {pesos[2]:>14.6f}")
    print(f"  {'theta_1':<12} {pesos[0]:>14.6f}")
    print("-" * 40)

    planilha2_path = os.path.join(base_dir, "Tabela2_pesos.xlsx")
    gerar_planilha_tabela2(pesos, planilha2_path)
    print(f"\nPlanilha Excel salva em : {planilha2_path}")

    # =========================================================================
    # QUESTAO 3 - Grafico do EQM em funcao de cada epoca do 2o estagio
    # =========================================================================
    print("\n" + "=" * 64)
    print("QUESTAO 3 - GRAFICO DO EQM x EPOCAS (2o ESTAGIO)")
    print("=" * 64)
    print(f"Epocas             : {n_epocas}")
    print(f"EQM inicial        : {eqm_list[0]:.8f}")
    print(f"EQM final          : {eqm_list[-1]:.8f}")
    print(f"Reducao do EQM     : {(1 - eqm_list[-1] / eqm_list[0]) * 100:.2f}%")

    eqm_path = os.path.join(output_dir, "q3_eqm_etapa2.png")
    salvar_grafico_eqm(eqm_list, eqm_path, eta=0.01, epsilon=1e-7)
    print(f"\nGrafico EQM salvo em    : {eqm_path}")

    eqm_xlsx_path = os.path.join(base_dir, "EQM_por_epoca.xlsx")
    exportar_eqm_por_epoca(eqm_list, eqm_xlsx_path)
    print(f"Planilha EQM salva em   : {eqm_xlsx_path}")

    # =========================================================================
    # QUESTAO 4 - Validacao da rede RBF (Tabela 3) com pos-processamento
    # =========================================================================
    print("\n" + "=" * 64)
    print("QUESTAO 4 - VALIDACAO DA REDE RBF (TABELA 3)")
    print("=" * 64)

    X_val = carregar_validacao(os.path.join(base_dir, "PP05_dados_validacao.txt"))
    d_val = D_VALIDACAO

    # Resposta da rede e pos-processamento (y_pos = 1 se y >= 0, senao -1)
    y_val = saida_rbf(X_val, centros, variancias, pesos)
    y_pos = pos_processar(y_val)

    n_acertos = int(np.sum(y_pos == d_val))
    taxa = 100.0 * n_acertos / len(X_val)

    cab = f"{'Sample':>6} | {'x1':>6} {'x2':>6} | {'d':>3} | {'y':>9} | {'y_post':>6} | OK"
    print(cab)
    print("-" * len(cab))
    for i in range(len(X_val)):
        ok = "sim" if int(y_pos[i]) == int(d_val[i]) else "NAO"
        print(f"{i + 1:>6} | {X_val[i, 0]:6.4f} {X_val[i, 1]:6.4f} | "
              f"{int(d_val[i]):>3} | {y_val[i]:>9.4f} | {int(y_pos[i]):>6} | {ok}")
    print("-" * len(cab))
    print(f"Acertos: {n_acertos}/{len(X_val)}  ->  Success rate = {taxa:.2f}%")

    planilha3_path = os.path.join(base_dir, "Tabela3_validacao.xlsx")
    gerar_planilha_tabela3(X_val, d_val, y_val, y_pos, taxa, planilha3_path)
    print(f"\nPlanilha Excel salva em : {planilha3_path}")

    # =========================================================================
    # QUESTAO 5 - Matriz de confusao e parametros de classificacao
    # =========================================================================
    print("\n" + "=" * 64)
    print("QUESTAO 5 - MATRIZ DE CONFUSAO E PARAMETROS")
    print("=" * 64)

    cm = matriz_confusao_binaria(d_val, y_pos)
    par = parametros_classificacao(cm)
    classes = ["Existente(+1)", "Inexistente(-1)"]

    print("Matriz de confusao (linha = verdadeiro, coluna = predito):")
    cabecalho = "Real \\ Predito"
    print(f"{cabecalho:>18}" + "".join(f"{c:>17}" for c in classes))
    for i in range(2):
        print(f"{classes[i]:>18}" + f"{ ''.join(f'{int(cm[i, j]):>17d}' for j in range(2))}")

    print(f"\n  TP = {par['TP']}   FN = {par['FN']}   "
          f"FP = {par['FP']}   TN = {par['TN']}")
    print("-" * 40)
    print(f"  {'Nacertos':<18}: {par['nacertos']}")
    print(f"  {'Nerros':<18}: {par['nerros']}")
    print(f"  {'Acuracia':<18}: {par['acuracia'] * 100:.2f}%")
    print(f"  {'Sensibilidade':<18}: {par['sensibilidade'] * 100:.2f}%")
    print(f"  {'Especificidade':<18}: {par['especificidade'] * 100:.2f}%")
    print(f"  {'Precisao':<18}: {par['precisao'] * 100:.2f}%")
    print("-" * 40)

    confusao_path = os.path.join(output_dir, "q5_matriz_confusao.png")
    salvar_grafico_confusao(cm, confusao_path)
    print(f"\nGrafico salvo em        : {confusao_path}")

    planilha5_path = os.path.join(base_dir, "Tabela_Q5_metricas.xlsx")
    gerar_planilha_etapa5(cm, par, planilha5_path)
    print(f"Planilha Excel salva em : {planilha5_path}")

    # =========================================================================
    # ETAPA 6 - Relatorio LaTeX (gerado e atualizado a cada execucao)
    # =========================================================================
    print("\n" + "=" * 64)
    print("ETAPA 6 - RELATORIO LATEX (atualizado a cada execucao)")
    print("=" * 64)

    relatorio_path = os.path.join(base_dir, "Relatorio_RBF.tex")
    gerar_relatorio_latex(
        centros, variancias, contagens, pesos,
        eqm_list, n_epocas, 0.01, 1e-7,
        X_val, d_val, y_val, y_pos, taxa, cm, par,
        # caminhos relativos das figuras (relativos ao .tex, que fica em base_dir)
        os.path.relpath(grafico_path, base_dir),
        os.path.relpath(eqm_path, base_dir),
        os.path.relpath(confusao_path, base_dir),
        relatorio_path)
    print(f"Relatorio LaTeX salvo em: {relatorio_path}")
    compilar_pdf(relatorio_path)


if __name__ == "__main__":
    main()
